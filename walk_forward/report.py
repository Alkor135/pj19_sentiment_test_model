from __future__ import annotations

from html import escape
from pathlib import Path
from typing import Any

import pandas as pd
import plotly.graph_objects as go
import plotly.io as pio
import typer
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


DEFAULT_RESULTS_DIR = Path(__file__).resolve().parent / "results"
DEFAULT_SUMMARY_CSV = DEFAULT_RESULTS_DIR / "summary.csv"
DEFAULT_OUTPUT_HTML = DEFAULT_RESULTS_DIR / "walk_forward_report.html"
DEFAULT_OUTPUT_XLSX = DEFAULT_RESULTS_DIR / "walk_forward_report.xlsx"
SUMMARY_COLUMNS = (
    "status",
    "ticker",
    "model_dir",
    "sentiment_model",
    "source_date",
    "trades",
    "pnl",
    "skip_reason",
    "error",
)
TRADE_COLUMNS = (
    "ticker",
    "model_dir",
    "sentiment_model",
    "source_date",
    "pnl",
    "direction",
    "action",
    "sentiment",
)
GROUP_KEYS = ["ticker", "model_dir", "sentiment_model"]

app = typer.Typer(help="Собрать Excel и HTML отчёты walk-forward бэктеста.")


def normalize_summary(summary: pd.DataFrame) -> pd.DataFrame:
    result = summary.copy()
    for column in SUMMARY_COLUMNS:
        if column not in result.columns:
            result[column] = ""

    result["status"] = result["status"].fillna("").astype(str)
    result["ticker"] = result["ticker"].fillna("").astype(str).str.upper()
    result["model_dir"] = result["model_dir"].fillna("").astype(str)
    result["sentiment_model"] = result["sentiment_model"].fillna("").astype(str)
    result["source_date"] = pd.to_datetime(result["source_date"], errors="coerce").dt.date
    result["trades"] = pd.to_numeric(result["trades"], errors="coerce").fillna(0).astype(int)
    result["pnl"] = pd.to_numeric(result["pnl"], errors="coerce").fillna(0.0).astype(float)
    result["skip_reason"] = result["skip_reason"].fillna("").astype(str)
    result["error"] = result["error"].fillna("").astype(str)
    return result


def normalize_trades(trades: pd.DataFrame) -> pd.DataFrame:
    result = trades.copy()
    for column in TRADE_COLUMNS:
        if column not in result.columns:
            result[column] = ""

    result["ticker"] = result["ticker"].fillna("").astype(str).str.upper()
    result["model_dir"] = result["model_dir"].fillna("").astype(str)
    result["sentiment_model"] = result["sentiment_model"].fillna("").astype(str)
    result["source_date"] = pd.to_datetime(result["source_date"], errors="coerce").dt.date
    result["pnl"] = pd.to_numeric(result["pnl"], errors="coerce").fillna(0.0).astype(float)
    result["sentiment"] = pd.to_numeric(result["sentiment"], errors="coerce")
    result["direction"] = result["direction"].fillna("").astype(str)
    result["action"] = result["action"].fillna("").astype(str)
    return result.dropna(subset=["source_date"])


def _empty_leaderboard() -> pd.DataFrame:
    return pd.DataFrame(
        columns=[
            "ticker",
            "rank",
            "model_dir",
            "sentiment_model",
            "days",
            "trades",
            "total_pnl",
            "winrate",
            "profit_factor",
            "max_drawdown",
            "recovery_factor",
            "avg_trade",
            "best_day",
            "worst_day",
            "score",
            "skipped_days",
            "error_days",
        ]
    )


def _metric_row(key: tuple[Any, ...], summary: pd.DataFrame, trades: pd.DataFrame) -> dict[str, Any]:
    ticker, model_dir, sentiment_model = key
    model_summary = summary[
        (summary["ticker"] == ticker)
        & (summary["model_dir"] == model_dir)
        & (summary["sentiment_model"] == sentiment_model)
    ]
    model_trades = trades[
        (trades["ticker"] == ticker)
        & (trades["model_dir"] == model_dir)
        & (trades["sentiment_model"] == sentiment_model)
    ].sort_values("source_date")

    daily_pnl = model_trades.groupby("source_date")["pnl"].sum().sort_index()
    trades_count = int(len(model_trades))
    total_pnl = float(model_trades["pnl"].sum()) if trades_count else 0.0
    gross_profit = float(model_trades.loc[model_trades["pnl"] > 0, "pnl"].sum()) if trades_count else 0.0
    gross_loss = abs(float(model_trades.loc[model_trades["pnl"] < 0, "pnl"].sum())) if trades_count else 0.0
    profit_factor = gross_profit / gross_loss if gross_loss else float("inf")
    winrate = float((model_trades["pnl"] > 0).mean() * 100) if trades_count else 0.0
    avg_trade = total_pnl / trades_count if trades_count else 0.0
    best_day = float(daily_pnl.max()) if not daily_pnl.empty else 0.0
    worst_day = float(daily_pnl.min()) if not daily_pnl.empty else 0.0
    max_drawdown = 0.0
    if not daily_pnl.empty:
        cum = daily_pnl.cumsum()
        max_drawdown = float((cum - cum.cummax()).min())
    recovery_factor = total_pnl / abs(max_drawdown) if max_drawdown else float("inf")
    score = total_pnl + max_drawdown * 0.5

    summary_days = int(model_summary["source_date"].nunique()) if not model_summary.empty else 0
    trade_days = int(daily_pnl.index.nunique()) if not daily_pnl.empty else 0
    return {
        "ticker": ticker,
        "rank": 0,
        "model_dir": model_dir,
        "sentiment_model": sentiment_model,
        "days": max(summary_days, trade_days),
        "trades": trades_count,
        "total_pnl": total_pnl,
        "winrate": winrate,
        "profit_factor": profit_factor,
        "max_drawdown": max_drawdown,
        "recovery_factor": recovery_factor,
        "avg_trade": avg_trade,
        "best_day": best_day,
        "worst_day": worst_day,
        "score": score,
        "skipped_days": int((model_summary["status"] == "skipped").sum()) if not model_summary.empty else 0,
        "error_days": int((model_summary["status"] == "error").sum()) if not model_summary.empty else 0,
    }


def build_leaderboard(summary: pd.DataFrame, trades: pd.DataFrame) -> pd.DataFrame:
    summary = normalize_summary(summary)
    trades = normalize_trades(trades)
    ok_summary = summary[summary["status"] == "ok"]

    keys: set[tuple[Any, ...]] = set()
    if not ok_summary.empty:
        keys.update(tuple(row) for row in ok_summary[GROUP_KEYS].drop_duplicates().to_numpy())
    if not trades.empty:
        keys.update(tuple(row) for row in trades[GROUP_KEYS].drop_duplicates().to_numpy())
    if not keys:
        return _empty_leaderboard()

    result = pd.DataFrame([_metric_row(key, summary, trades) for key in sorted(keys)])
    result = result.sort_values(["ticker", "score", "total_pnl"], ascending=[True, False, False])
    result["rank"] = result.groupby("ticker").cumcount() + 1
    return result[_empty_leaderboard().columns].reset_index(drop=True)


def build_ticker_summary(leaderboard: pd.DataFrame) -> pd.DataFrame:
    if leaderboard.empty:
        return pd.DataFrame(
            columns=[
                "ticker",
                "models",
                "best_model",
                "best_score",
                "total_pnl",
                "avg_model_pnl",
                "trades",
                "best_winrate",
                "worst_drawdown",
            ]
        )

    rows: list[dict[str, Any]] = []
    for ticker, group in leaderboard.sort_values(["ticker", "rank"]).groupby("ticker", sort=True):
        best = group.iloc[0]
        rows.append(
            {
                "ticker": ticker,
                "models": int(group["model_dir"].nunique()),
                "best_model": best["model_dir"],
                "best_score": float(best["score"]),
                "total_pnl": float(group["total_pnl"].sum()),
                "avg_model_pnl": float(group["total_pnl"].mean()),
                "trades": int(group["trades"].sum()),
                "best_winrate": float(group["winrate"].max()),
                "worst_drawdown": float(group["max_drawdown"].min()),
            }
        )
    return pd.DataFrame(rows)


def _empty_errors() -> pd.DataFrame:
    return pd.DataFrame(columns=["ticker", "model_dir", "sentiment_model", "status", "error"])


def _resolve_ticker_dir(results_dir: Path, ticker: str) -> Path:
    direct = results_dir / ticker
    if direct.exists():
        return direct

    ticker_lc = ticker.lower()
    if results_dir.exists():
        for child in results_dir.iterdir():
            if child.is_dir() and child.name.lower() == ticker_lc:
                return child
    return direct


def load_all_trades(results_dir: Path, summary: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    summary = normalize_summary(summary)
    ok_summary = summary[summary["status"] == "ok"]
    frames: list[pd.DataFrame] = []
    errors: list[dict[str, str]] = []

    for item in ok_summary[GROUP_KEYS].drop_duplicates().to_dict("records"):
        ticker = str(item["ticker"])
        model_dir = str(item["model_dir"])
        sentiment_model = str(item["sentiment_model"])
        trades_path = _resolve_ticker_dir(results_dir, ticker) / model_dir / "trades.csv"

        if not trades_path.exists():
            errors.append(
                {
                    "ticker": ticker,
                    "model_dir": model_dir,
                    "sentiment_model": sentiment_model,
                    "status": "error",
                    "error": f"Не найден файл trades.csv: {trades_path}",
                }
            )
            continue

        try:
            frame = pd.read_csv(trades_path, encoding="utf-8-sig")
        except Exception as exc:
            errors.append(
                {
                    "ticker": ticker,
                    "model_dir": model_dir,
                    "sentiment_model": sentiment_model,
                    "status": "error",
                    "error": f"Не удалось прочитать trades.csv: {exc}",
                }
            )
            continue

        frame["ticker"] = ticker
        frame["model_dir"] = model_dir
        frame["sentiment_model"] = sentiment_model
        frames.append(frame)

    trades = normalize_trades(pd.concat(frames, ignore_index=True)) if frames else normalize_trades(pd.DataFrame())
    error_frame = pd.DataFrame(errors) if errors else _empty_errors()
    return trades, error_frame


def _build_period_matrix(trades: pd.DataFrame, period_column: str) -> pd.DataFrame:
    trades = normalize_trades(trades)
    if trades.empty:
        return pd.DataFrame()

    result = trades.copy()
    result["series"] = result["ticker"] + " / " + result["model_dir"]
    if period_column == "month":
        result["period"] = pd.to_datetime(result["source_date"]).dt.to_period("M").astype(str)
    else:
        result["period"] = pd.to_datetime(result["source_date"]).dt.strftime("%Y-%m-%d")

    matrix = result.pivot_table(
        index="series",
        columns="period",
        values="pnl",
        aggfunc="sum",
        fill_value=0.0,
    )
    return matrix.sort_index().sort_index(axis=1)


def build_monthly_matrix(trades: pd.DataFrame) -> pd.DataFrame:
    return _build_period_matrix(trades, "month")


def build_daily_matrix(trades: pd.DataFrame) -> pd.DataFrame:
    return _build_period_matrix(trades, "day")


def build_dashboard(
    summary: pd.DataFrame,
    trades: pd.DataFrame,
    leaderboard: pd.DataFrame,
    ticker_summary: pd.DataFrame,
    errors: pd.DataFrame,
) -> pd.DataFrame:
    summary = normalize_summary(summary)
    trades = normalize_trades(trades)
    total_pnl = float(trades["pnl"].sum()) if not trades.empty else 0.0
    best = leaderboard.iloc[0] if not leaderboard.empty else None
    return pd.DataFrame(
        [
            {"Показатель": "Тикеров", "Значение": int(ticker_summary["ticker"].nunique()) if not ticker_summary.empty else 0},
            {"Показатель": "Моделей", "Значение": int(leaderboard["model_dir"].nunique()) if not leaderboard.empty else 0},
            {"Показатель": "Дней в summary", "Значение": int(summary["source_date"].nunique()) if not summary.empty else 0},
            {"Показатель": "Сделок", "Значение": int(len(trades))},
            {"Показатель": "Total P/L", "Значение": total_pnl},
            {"Показатель": "Лучшая модель", "Значение": "" if best is None else f"{best['ticker']} / {best['model_dir']}"},
            {"Показатель": "Ошибок", "Значение": int(len(errors))},
        ]
    )


def _excel_safe(frame: pd.DataFrame) -> pd.DataFrame:
    return frame.copy()


def _sheet_name(value: str, used: set[str]) -> str:
    base = "".join(ch for ch in value if ch not in r"[]:*?/\\")[:31] or "Sheet"
    name = base
    suffix = 2
    while name in used:
        tail = f"_{suffix}"
        name = f"{base[:31 - len(tail)]}{tail}"
        suffix += 1
    used.add(name)
    return name


def write_excel_report(
    *,
    summary: pd.DataFrame,
    trades: pd.DataFrame,
    leaderboard: pd.DataFrame,
    ticker_summary: pd.DataFrame,
    monthly_matrix: pd.DataFrame,
    daily_matrix: pd.DataFrame,
    errors: pd.DataFrame,
    output_xlsx: Path,
) -> None:
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    summary = normalize_summary(summary)
    trades = normalize_trades(trades)
    errors = errors.copy() if not errors.empty else _empty_errors()
    dashboard = build_dashboard(summary, trades, leaderboard, ticker_summary, errors)
    used_sheets = {
        "Dashboard",
        "Leaderboard",
        "Ticker_Summary",
        "Monthly_Matrix",
        "Daily_Matrix",
        "Raw_Summary",
        "Raw_Trades",
        "Errors",
    }

    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        _excel_safe(dashboard).to_excel(writer, sheet_name="Dashboard", index=False, inf_rep="inf")
        _excel_safe(leaderboard).to_excel(writer, sheet_name="Leaderboard", index=False, inf_rep="inf")
        _excel_safe(ticker_summary).to_excel(writer, sheet_name="Ticker_Summary", index=False, inf_rep="inf")
        _excel_safe(monthly_matrix).to_excel(writer, sheet_name="Monthly_Matrix", inf_rep="inf")
        _excel_safe(daily_matrix).to_excel(writer, sheet_name="Daily_Matrix", inf_rep="inf")

        tickers = sorted(
            set(leaderboard["ticker"].dropna())
            | set(ticker_summary["ticker"].dropna())
            | set(trades["ticker"].dropna())
        )
        for ticker in tickers:
            ticker_trades = trades[trades["ticker"] == ticker]
            ticker_leaderboard = leaderboard[leaderboard["ticker"] == ticker]
            sheet = _sheet_name(str(ticker), used_sheets)
            if not ticker_trades.empty:
                frame = ticker_trades
            else:
                frame = ticker_leaderboard
            _excel_safe(frame).to_excel(writer, sheet_name=sheet, index=False, inf_rep="inf")

        _excel_safe(summary).to_excel(writer, sheet_name="Raw_Summary", index=False, inf_rep="inf")
        _excel_safe(trades).to_excel(writer, sheet_name="Raw_Trades", index=False, inf_rep="inf")
        _excel_safe(errors).to_excel(writer, sheet_name="Errors", index=False, inf_rep="inf")

    _style_workbook(output_xlsx)


def _style_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    positive_fill = PatternFill("solid", fgColor="E2F0D9")
    negative_fill = PatternFill("solid", fgColor="FCE4D6")

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.showGridLines = False
        if worksheet.max_row > 1 and worksheet.max_column > 0:
            worksheet.auto_filter.ref = worksheet.dimensions

        headers: dict[int, str] = {}
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            headers[cell.column] = str(cell.value or "").lower()

        for column_cells in worksheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 44)
            worksheet.column_dimensions[column_cells[0].column_letter].width = max(width, 10)

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                header = headers.get(cell.column, "")
                if isinstance(cell.value, (int, float)):
                    if any(token in header for token in ("pnl", "p/l", "score", "drawdown", "profit", "avg", "best", "worst")):
                        cell.number_format = "#,##0.00"
                    if any(token in header for token in ("winrate", "rate")):
                        cell.number_format = "0.00"
                    if any(token in header for token in ("pnl", "p/l", "score")):
                        if cell.value > 0:
                            cell.fill = positive_fill
                        elif cell.value < 0:
                            cell.fill = negative_fill

    workbook.save(path)


def _table_html(frame: pd.DataFrame, max_rows: int = 50, index: bool = False) -> str:
    if frame.empty:
        return '<p class="muted">Нет данных</p>'
    visible = frame.head(max_rows).copy()
    return visible.to_html(index=index, classes="data-table", border=0, escape=True)


def _kpi_html(label: str, value: Any) -> str:
    return f'<div class="kpi"><span>{escape(label)}</span><strong>{escape(str(value))}</strong></div>'


def _format_float(value: Any) -> str:
    if isinstance(value, (int, float)):
        if value == float("inf"):
            return "inf"
        if value == float("-inf"):
            return "-inf"
        return f"{value:,.2f}"
    return str(value)


def _plot_div(fig: go.Figure, *, include_plotlyjs: bool) -> str:
    return pio.to_html(
        fig,
        full_html=False,
        include_plotlyjs=include_plotlyjs,
        config={"displayModeBar": False, "responsive": True},
    )


def _overall_daily_pnl(trades: pd.DataFrame) -> pd.Series:
    trades = normalize_trades(trades)
    if trades.empty:
        return pd.Series(dtype=float)
    return trades.groupby("source_date")["pnl"].sum().sort_index()


def _equity_figure(trades: pd.DataFrame) -> go.Figure:
    daily = _overall_daily_pnl(trades)
    fig = go.Figure()
    if not daily.empty:
        fig.add_trace(
            go.Scatter(
                x=[str(item) for item in daily.index],
                y=daily.cumsum(),
                mode="lines",
                name="Equity",
                line={"color": "#2563eb", "width": 2},
            )
        )
    fig.update_layout(title="Equity Curve", template="plotly_white", height=360, margin={"l": 40, "r": 20, "t": 55, "b": 40})
    return fig


def _drawdown_figure(trades: pd.DataFrame) -> go.Figure:
    daily = _overall_daily_pnl(trades)
    fig = go.Figure()
    if not daily.empty:
        cum = daily.cumsum()
        drawdown = cum - cum.cummax()
        fig.add_trace(
            go.Scatter(
                x=[str(item) for item in drawdown.index],
                y=drawdown,
                fill="tozeroy",
                mode="lines",
                name="Drawdown",
                line={"color": "#dc2626", "width": 2},
            )
        )
    fig.update_layout(title="Drawdown", template="plotly_white", height=320, margin={"l": 40, "r": 20, "t": 55, "b": 40})
    return fig


def _daily_bar_figure(trades: pd.DataFrame) -> go.Figure:
    daily = _overall_daily_pnl(trades)
    fig = go.Figure()
    if not daily.empty:
        colors = ["#16a34a" if value >= 0 else "#dc2626" for value in daily]
        fig.add_trace(go.Bar(x=[str(item) for item in daily.index], y=daily, marker={"color": colors}, name="Daily P/L"))
    fig.update_layout(title="Daily P/L", template="plotly_white", height=320, margin={"l": 40, "r": 20, "t": 55, "b": 40})
    return fig


def _monthly_heatmap_figure(monthly_matrix: pd.DataFrame) -> go.Figure:
    fig = go.Figure()
    if not monthly_matrix.empty:
        fig.add_trace(
            go.Heatmap(
                z=monthly_matrix.values,
                x=list(monthly_matrix.columns),
                y=list(monthly_matrix.index),
                colorscale="RdYlGn",
                zmid=0,
                colorbar={"title": "P/L"},
            )
        )
    fig.update_layout(title="Monthly Matrix", template="plotly_white", height=420, margin={"l": 120, "r": 20, "t": 55, "b": 50})
    return fig


def _ticker_equity_figure(ticker: str, trades: pd.DataFrame, leaderboard: pd.DataFrame) -> go.Figure:
    fig = go.Figure()
    top_models = leaderboard[(leaderboard["ticker"] == ticker) & (leaderboard["rank"] <= 5)]["model_dir"].tolist()
    ticker_trades = normalize_trades(trades)
    for model_dir in top_models:
        model_trades = ticker_trades[
            (ticker_trades["ticker"] == ticker)
            & (ticker_trades["model_dir"] == model_dir)
        ]
        if model_trades.empty:
            continue
        daily = model_trades.groupby("source_date")["pnl"].sum().sort_index()
        fig.add_trace(
            go.Scatter(
                x=[str(item) for item in daily.index],
                y=daily.cumsum(),
                mode="lines",
                name=model_dir,
            )
        )
    fig.update_layout(
        title=f"{ticker}: equity top-5 моделей",
        template="plotly_white",
        height=320,
        margin={"l": 40, "r": 20, "t": 55, "b": 40},
    )
    return fig


def build_html(
    *,
    summary: pd.DataFrame,
    trades: pd.DataFrame,
    leaderboard: pd.DataFrame,
    ticker_summary: pd.DataFrame,
    monthly_matrix: pd.DataFrame,
    daily_matrix: pd.DataFrame,
    errors: pd.DataFrame,
) -> str:
    summary = normalize_summary(summary)
    trades = normalize_trades(trades)
    total_pnl = float(trades["pnl"].sum()) if not trades.empty else 0.0
    max_drawdown = float(leaderboard["max_drawdown"].min()) if not leaderboard.empty else 0.0
    best_model = ""
    if not leaderboard.empty:
        best = leaderboard.iloc[0]
        best_model = f"{best['ticker']} / {best['model_dir']}"

    chart_parts: list[str] = [
        _plot_div(_equity_figure(trades), include_plotlyjs=True),
        _plot_div(_drawdown_figure(trades), include_plotlyjs=False),
        _plot_div(_daily_bar_figure(trades), include_plotlyjs=False),
        _plot_div(_monthly_heatmap_figure(monthly_matrix), include_plotlyjs=False),
    ]

    ticker_sections: list[str] = []
    for ticker in sorted(leaderboard["ticker"].unique()) if not leaderboard.empty else []:
        ticker_leaderboard = leaderboard[leaderboard["ticker"] == ticker].sort_values("rank")
        ticker_sections.append(
            f"""
            <section>
              <h2>{escape(str(ticker))}</h2>
              {_plot_div(_ticker_equity_figure(str(ticker), trades, leaderboard), include_plotlyjs=False)}
              {_table_html(ticker_leaderboard, max_rows=30)}
            </section>
            """
        )

    return f"""<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Walk-Forward Dashboard</title>
  <style>
    :root {{
      color-scheme: light;
      font-family: "Segoe UI", Arial, sans-serif;
      color: #172033;
      background: #f6f7f9;
    }}
    body {{ margin: 0; }}
    main {{ max-width: 1280px; margin: 0 auto; padding: 28px; }}
    h1 {{ font-size: 34px; margin: 0 0 18px; letter-spacing: 0; }}
    h2 {{ font-size: 22px; margin: 32px 0 14px; letter-spacing: 0; }}
    .kpis {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 12px; }}
    .kpi {{ background: #ffffff; border: 1px solid #dde2ea; border-radius: 8px; padding: 14px 16px; }}
    .kpi span {{ display: block; color: #647084; font-size: 13px; margin-bottom: 8px; }}
    .kpi strong {{ display: block; font-size: 20px; overflow-wrap: anywhere; }}
    section {{ margin-top: 22px; }}
    .panel {{ background: #ffffff; border: 1px solid #dde2ea; border-radius: 8px; padding: 16px; margin-top: 12px; overflow-x: auto; }}
    .grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(420px, 1fr)); gap: 14px; }}
    .data-table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
    .data-table th {{ text-align: left; background: #22324a; color: #ffffff; padding: 8px; white-space: nowrap; }}
    .data-table td {{ border-bottom: 1px solid #e6e9ef; padding: 7px 8px; white-space: nowrap; }}
    .muted {{ color: #647084; }}
  </style>
</head>
<body>
<main>
  <h1>Walk-Forward Dashboard</h1>
  <div class="kpis">
    {_kpi_html("Total P/L", _format_float(total_pnl))}
    {_kpi_html("Сделок", len(trades))}
    {_kpi_html("Моделей", leaderboard["model_dir"].nunique() if not leaderboard.empty else 0)}
    {_kpi_html("Max drawdown", _format_float(max_drawdown))}
    {_kpi_html("Лучшая модель", best_model)}
    {_kpi_html("Ошибок", len(errors))}
  </div>

  <section>
    <h2>Графики</h2>
    <div class="grid">
      {"".join(f'<div class="panel">{part}</div>' for part in chart_parts)}
    </div>
  </section>

  <section>
    <h2>Лучшие модели по тикерам</h2>
    <div class="panel">{_table_html(ticker_summary, max_rows=50)}</div>
    <div class="panel">{_table_html(leaderboard, max_rows=100)}</div>
  </section>

  {"".join(ticker_sections)}

  <section>
    <h2>Ошибки и пропуски</h2>
    <div class="panel">{_table_html(errors, max_rows=100)}</div>
  </section>
</main>
</body>
</html>
"""


def _summary_error_rows(summary: pd.DataFrame) -> pd.DataFrame:
    summary = normalize_summary(summary)
    errors = summary[summary["status"] != "ok"].copy()
    if errors.empty:
        return _empty_errors()
    errors["error"] = errors.apply(
        lambda row: row["error"] or row["skip_reason"] or row["status"],
        axis=1,
    )
    return errors[["ticker", "model_dir", "sentiment_model", "status", "error"]]


def build_report(
    *,
    summary_csv: Path = DEFAULT_SUMMARY_CSV,
    results_dir: Path = DEFAULT_RESULTS_DIR,
    output_html: Path = DEFAULT_OUTPUT_HTML,
    output_xlsx: Path = DEFAULT_OUTPUT_XLSX,
) -> tuple[Path, Path]:
    if not summary_csv.exists():
        raise FileNotFoundError(f"summary.csv не найден: {summary_csv}")

    summary = pd.read_csv(summary_csv, encoding="utf-8-sig")
    trades, load_errors = load_all_trades(results_dir, summary)
    summary = normalize_summary(summary)
    errors = pd.concat([_summary_error_rows(summary), load_errors], ignore_index=True)
    leaderboard = build_leaderboard(summary, trades)
    ticker_summary = build_ticker_summary(leaderboard)
    monthly_matrix = build_monthly_matrix(trades)
    daily_matrix = build_daily_matrix(trades)

    write_excel_report(
        summary=summary,
        trades=trades,
        leaderboard=leaderboard,
        ticker_summary=ticker_summary,
        monthly_matrix=monthly_matrix,
        daily_matrix=daily_matrix,
        errors=errors,
        output_xlsx=output_xlsx,
    )

    output_html.parent.mkdir(parents=True, exist_ok=True)
    output_html.write_text(
        build_html(
            summary=summary,
            trades=trades,
            leaderboard=leaderboard,
            ticker_summary=ticker_summary,
            monthly_matrix=monthly_matrix,
            daily_matrix=daily_matrix,
            errors=errors,
        ),
        encoding="utf-8",
    )
    return output_html, output_xlsx


@app.command()
def main(
    summary_csv: Path = typer.Option(DEFAULT_SUMMARY_CSV, "--summary-csv", help="Путь к summary.csv."),
    results_dir: Path = typer.Option(DEFAULT_RESULTS_DIR, "--results-dir", help="Папка результатов walk-forward."),
    output_html: Path = typer.Option(DEFAULT_OUTPUT_HTML, "--output-html", help="HTML отчёт."),
    output_xlsx: Path = typer.Option(DEFAULT_OUTPUT_XLSX, "--output-xlsx", help="Excel отчёт."),
) -> None:
    html_path, xlsx_path = build_report(
        summary_csv=summary_csv,
        results_dir=results_dir,
        output_html=output_html,
        output_xlsx=output_xlsx,
    )
    typer.echo(f"HTML отчёт: {html_path}")
    typer.echo(f"Excel отчёт: {xlsx_path}")


if __name__ == "__main__":
    app()
