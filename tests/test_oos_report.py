from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from oos import report


def _sample_summary() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "status": "ok",
                "ticker": "RTS",
                "model_dir": "model_a",
                "sentiment_model": "model:a",
                "month": "2025-10",
                "train_rows": 10,
                "test_rows": 2,
                "trades": 2,
                "total_pnl": 10,
                "winrate": 50,
                "max_drawdown": -3,
                "error": "",
            },
            {
                "status": "ok",
                "ticker": "RTS",
                "model_dir": "model_a",
                "sentiment_model": "model:a",
                "month": "2025-11",
                "train_rows": 10,
                "test_rows": 2,
                "trades": 1,
                "total_pnl": -5,
                "winrate": 0,
                "max_drawdown": -7,
                "error": "",
            },
            {
                "status": "ok",
                "ticker": "RTS",
                "model_dir": "model_b",
                "sentiment_model": "model:b",
                "month": "2025-10",
                "train_rows": 10,
                "test_rows": 2,
                "trades": 2,
                "total_pnl": 20,
                "winrate": 100,
                "max_drawdown": -2,
                "error": "",
            },
            {
                "status": "ok",
                "ticker": "RTS",
                "model_dir": "model_b",
                "sentiment_model": "model:b",
                "month": "2025-11",
                "train_rows": 10,
                "test_rows": 2,
                "trades": 2,
                "total_pnl": 15,
                "winrate": 50,
                "max_drawdown": -1,
                "error": "",
            },
            {
                "status": "error",
                "ticker": "MIX",
                "model_dir": "missing_model",
                "sentiment_model": "missing:model",
                "month": "2025-10",
                "error": "sentiment PKL не найден",
            },
        ]
    )


def test_build_leaderboard_aggregates_months_and_scores_models() -> None:
    leaderboard = report.build_leaderboard(_sample_summary())

    rts = leaderboard[leaderboard["ticker"] == "RTS"].reset_index(drop=True)

    assert rts["model_dir"].tolist() == ["model_b", "model_a"]
    assert rts.loc[0, "months"] == 2
    assert rts.loc[0, "total_pnl"] == 35
    assert rts.loc[0, "median_pnl_month"] == 17.5
    assert rts.loc[0, "win_months_pct"] == 100.0
    assert rts.loc[0, "pnl_per_trade"] == 8.75
    assert rts.loc[0, "worst_month"] == "2025-11"
    assert rts.loc[0, "max_drawdown"] == -2


def test_build_monthly_matrix_keeps_models_as_rows_and_months_as_columns() -> None:
    matrix = report.build_monthly_matrix(_sample_summary())

    assert matrix.loc["RTS / model_a", "2025-10"] == 10
    assert matrix.loc["RTS / model_a", "2025-11"] == -5
    assert matrix.loc["RTS / model_b", "2025-10"] == 20
    assert matrix.loc["RTS / model_b", "2025-11"] == 15


def test_build_report_writes_html_and_excel_workbook(tmp_path: Path) -> None:
    summary_csv = tmp_path / "summary.csv"
    output_html = tmp_path / "oos_report.html"
    output_xlsx = tmp_path / "oos_report.xlsx"
    _sample_summary().to_csv(summary_csv, index=False, encoding="utf-8-sig")

    report.build_report(
        summary_csv=summary_csv,
        output_html=output_html,
        output_xlsx=output_xlsx,
    )

    html = output_html.read_text(encoding="utf-8")
    assert "OOS Dashboard" in html
    assert "RTS / model_b" in html
    assert "Ошибки и пропуски" in html

    workbook = load_workbook(output_xlsx, read_only=True)
    assert workbook.sheetnames == [
        "Dashboard",
        "Leaderboard",
        "Monthly_Matrix",
        "RTS",
        "Raw_Summary",
        "Errors",
    ]
