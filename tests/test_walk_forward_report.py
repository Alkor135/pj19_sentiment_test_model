from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from walk_forward import report


def _sample_summary() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "status": "ok",
                "ticker": "RTS",
                "model_dir": "model_a",
                "sentiment_model": "model:a",
                "source_date": "2026-04-01",
                "trades": 1,
                "pnl": 10,
                "skip_reason": "",
                "error": "",
            },
            {
                "status": "ok",
                "ticker": "RTS",
                "model_dir": "model_a",
                "sentiment_model": "model:a",
                "source_date": "2026-04-02",
                "trades": 1,
                "pnl": -4,
                "skip_reason": "",
                "error": "",
            },
            {
                "status": "ok",
                "ticker": "RTS",
                "model_dir": "model_b",
                "sentiment_model": "model:b",
                "source_date": "2026-04-01",
                "trades": 1,
                "pnl": 20,
                "skip_reason": "",
                "error": "",
            },
            {
                "status": "skipped",
                "ticker": "MIX",
                "model_dir": "model_c",
                "sentiment_model": "model:c",
                "source_date": "2026-04-01",
                "trades": 0,
                "pnl": 0,
                "skip_reason": "insufficient_train_rows",
                "error": "",
            },
        ]
    )


def _sample_trades() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "ticker": "RTS",
                "model_dir": "model_a",
                "sentiment_model": "model:a",
                "source_date": date(2026, 4, 1),
                "pnl": 10.0,
                "direction": "LONG",
                "action": "follow",
                "sentiment": 1.0,
            },
            {
                "ticker": "RTS",
                "model_dir": "model_a",
                "sentiment_model": "model:a",
                "source_date": date(2026, 4, 2),
                "pnl": -4.0,
                "direction": "SHORT",
                "action": "invert",
                "sentiment": -1.0,
            },
            {
                "ticker": "RTS",
                "model_dir": "model_b",
                "sentiment_model": "model:b",
                "source_date": date(2026, 4, 1),
                "pnl": 20.0,
                "direction": "LONG",
                "action": "follow",
                "sentiment": 2.0,
            },
            {
                "ticker": "RTS",
                "model_dir": "model_b",
                "sentiment_model": "model:b",
                "source_date": date(2026, 4, 2),
                "pnl": 5.0,
                "direction": "LONG",
                "action": "follow",
                "sentiment": 2.0,
            },
        ]
    )


def test_build_leaderboard_scores_models_by_ticker() -> None:
    leaderboard = report.build_leaderboard(_sample_summary(), _sample_trades())
    rts = leaderboard[leaderboard["ticker"] == "RTS"].reset_index(drop=True)

    assert rts["model_dir"].tolist() == ["model_b", "model_a"]
    assert rts.loc[0, "rank"] == 1
    assert rts.loc[0, "trades"] == 2
    assert rts.loc[0, "total_pnl"] == 25.0
    assert rts.loc[0, "winrate"] == 100.0
    assert rts.loc[1, "profit_factor"] == 2.5


def test_build_ticker_summary_selects_best_model() -> None:
    leaderboard = report.build_leaderboard(_sample_summary(), _sample_trades())
    ticker_summary = report.build_ticker_summary(leaderboard)

    row = ticker_summary[ticker_summary["ticker"] == "RTS"].iloc[0]
    assert row["models"] == 2
    assert row["best_model"] == "model_b"
    assert row["total_pnl"] == 31.0


def test_build_monthly_and_daily_matrices() -> None:
    trades = _sample_trades()

    monthly = report.build_monthly_matrix(trades)
    daily = report.build_daily_matrix(trades)

    assert monthly.loc["RTS / model_a", "2026-04"] == 6.0
    assert monthly.loc["RTS / model_b", "2026-04"] == 25.0
    assert daily.loc["RTS / model_a", "2026-04-01"] == 10.0
    assert daily.loc["RTS / model_a", "2026-04-02"] == -4.0


def test_load_all_trades_records_missing_trade_file(tmp_path: Path) -> None:
    model_dir = tmp_path / "RTS" / "model_a"
    model_dir.mkdir(parents=True)
    pd.DataFrame(
        [
            {
                "ticker": "RTS",
                "model_dir": "model_a",
                "sentiment_model": "model:a",
                "source_date": "2026-04-01",
                "pnl": 1,
            }
        ]
    ).to_csv(model_dir / "trades.csv", index=False)
    summary = pd.DataFrame(
        [
            {"ticker": "RTS", "model_dir": "model_a", "sentiment_model": "model:a", "status": "ok"},
            {"ticker": "RTS", "model_dir": "missing", "sentiment_model": "missing", "status": "ok"},
        ]
    )

    trades, errors = report.load_all_trades(tmp_path, summary)

    assert len(trades) == 1
    assert errors.iloc[0]["model_dir"] == "missing"
    assert "trades.csv" in errors.iloc[0]["error"]


def test_write_excel_report_creates_expected_sheets(tmp_path: Path) -> None:
    output_xlsx = tmp_path / "walk_forward_report.xlsx"
    summary = _sample_summary()
    trades = _sample_trades()
    leaderboard = report.build_leaderboard(summary, trades)
    ticker_summary = report.build_ticker_summary(leaderboard)
    monthly = report.build_monthly_matrix(trades)
    daily = report.build_daily_matrix(trades)

    report.write_excel_report(
        summary=summary,
        trades=trades,
        leaderboard=leaderboard,
        ticker_summary=ticker_summary,
        monthly_matrix=monthly,
        daily_matrix=daily,
        errors=pd.DataFrame(),
        output_xlsx=output_xlsx,
    )

    workbook = load_workbook(output_xlsx, read_only=True)
    assert workbook.sheetnames == [
        "Dashboard",
        "Leaderboard",
        "Ticker_Summary",
        "Monthly_Matrix",
        "Daily_Matrix",
        "RTS",
        "Raw_Summary",
        "Raw_Trades",
        "Errors",
    ]


def test_build_report_writes_html_and_excel(tmp_path: Path) -> None:
    results_dir = tmp_path / "results"
    model_dir = results_dir / "RTS" / "model_a"
    model_dir.mkdir(parents=True)
    summary_csv = results_dir / "summary.csv"
    output_html = results_dir / "walk_forward_report.html"
    output_xlsx = results_dir / "walk_forward_report.xlsx"

    _sample_summary().to_csv(summary_csv, index=False, encoding="utf-8-sig")
    _sample_trades().to_csv(model_dir / "trades.csv", index=False, encoding="utf-8-sig")

    report.build_report(
        summary_csv=summary_csv,
        results_dir=results_dir,
        output_html=output_html,
        output_xlsx=output_xlsx,
    )

    html = output_html.read_text(encoding="utf-8")
    assert "Walk-Forward Dashboard" in html
    assert "Лучшие модели по тикерам" in html
    assert "RTS" in html
    assert output_xlsx.exists()
