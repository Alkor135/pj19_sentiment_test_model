from __future__ import annotations

from html import escape
from pathlib import Path
import sys
from typing import Any

import pandas as pd
import typer
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent.parent
if __package__ in {None, ""}:
    sys.path.insert(0, str(ROOT))


NUMERIC_COLUMNS = (
    "train_rows",
    "test_rows",
    "trades",
    "total_pnl",
    "winrate",
    "max_drawdown",
)
LEADERBOARD_COLUMNS = [
    "rank",
    "ticker",
    "model_dir",
    "sentiment_model",
    "months",
    "total_pnl",
    "avg_pnl_month",
    "median_pnl_month",
    "win_months",
    "win_months_pct",
    "trades",
    "pnl_per_trade",
    "avg_trade_winrate",
    "max_drawdown",
    "best_month",
    "best_month_pnl",
    "worst_month",
    "worst_month_pnl",
    "score",
]

app = typer.Typer(help="Строит HTML и Excel dashboard по OOS summary.csv.")


def normalize_summary(summary: pd.DataFrame) -> pd.DataFrame:
    out = summary.copy()
    for column in ("status", "ticker", "model_dir", "sentiment_model", "month", "error"):
        if column not in out.columns:
            out[column] = ""
        out[column] = out[column].fillna("").astype(str)
    out["ticker"] = out["ticker"].str.upper()
    for column in NUMERIC_COLUMNS:
        if column not in out.columns:
            out[column] = 0.0
        out[column] = pd.to_numeric(out[column], errors="coerce")
    return out


def ok_rows(summary: pd.DataFrame) -> pd.DataFrame:
    out = normalize_summary(summary)
    return out[out["status"] == "ok"].copy()


def error_rows(summary: pd.DataFrame) -> pd.DataFrame:
    out = normalize_summary(summary)
    return out[out["status"] != "ok"].copy()


def _weighted_average(values: pd.Series, weights: pd.Series) -> float:
    weights = weights.fillna(0).astype(float)
    total_weight = float(weights.sum())
    if total_weight == 0:
        return 0.0
    return float((values.fillna(0).astype(float) * weights).sum() / total_weight)


def build_leaderboard(summary: pd.DataFrame) -> pd.DataFrame:
    ok = ok_rows(summary)
    if ok.empty:
        return pd.DataFrame(columns=LEADERBOARD_COLUMNS)

    rows: list[dict[str, Any]] = []
    for (ticker, model_dir, sentiment_model), group in ok.groupby(
        ["ticker", "model_dir", "sentiment_model"],
        sort=True,
        dropna=False,
    ):
        group = group.sort_values("month").copy()
        total_pnl = float(group["total_pnl"].sum())
        trades = int(group["trades"].fillna(0).sum())
        months = int(len(group))
        best = group.loc[group["total_pnl"].idxmax()]
        worst = group.loc[group["total_pnl"].idxmin()]
        max_drawdown = float(group["max_drawdown"].min())
        win_months = int((group["total_pnl"] > 0).sum())
        win_months_pct = float(win_months / months * 100) if months else 0.0
        median_pnl = float(group["total_pnl"].median())
        score = (
            total_pnl
            + median_pnl * 2
            + win_months_pct * 100
            - abs(max_drawdown) * 0.5
        )
        rows.append(
            {
                "ticker": ticker,
                "model_dir": model_dir,
                "sentiment_model": sentiment_model,
                "months": months,
                "total_pnl": total_pnl,
                "avg_pnl_month": float(group["total_pnl"].mean()),
                "median_pnl_month": median_pnl,
                "win_months": win_months,
                "win_months_pct": win_months_pct,
                "trades": trades,
                "pnl_per_trade": float(total_pnl / trades) if trades else 0.0,
                "avg_trade_winrate": _weighted_average(group["winrate"], group["trades"]),
                "max_drawdown": max_drawdown,
                "best_month": str(best["month"]),
                "best_month_pnl": float(best["total_pnl"]),
                "worst_month": str(worst["month"]),
                "worst_month_pnl": float(worst["total_pnl"]),
                "score": float(score),
            }
        )

    leaderboard = pd.DataFrame(rows)
    leaderboard = leaderboard.sort_values(
        ["ticker", "score", "total_pnl"],
        ascending=[True, False, False],
    ).reset_index(drop=True)
    leaderboard.insert(
        0,
        "rank",
        leaderboard.groupby("ticker").cumcount() + 1,
    )
    return leaderboard[LEADERBOARD_COLUMNS]


def build_monthly_matrix(summary: pd.DataFrame) -> pd.DataFrame:
    ok = ok_rows(summary)
    if ok.empty:
        return pd.DataFrame()
    ok["row_label"] = ok["ticker"] + " / " + ok["model_dir"]
    matrix = ok.pivot_table(
        index="row_label",
        columns="month",
        values="total_pnl",
        aggfunc="sum",
    )
    return matrix.sort_index().sort_index(axis=1)


def _format_number(value: Any) -> str:
    if pd.isna(value):
        return ""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return escape(str(value))
    if abs(number) >= 100:
        return f"{number:,.0f}".replace(",", " ")
    return f"{number:,.2f}".replace(",", " ")


def _format_pct(value: Any) -> str:
    if pd.isna(value):
        return ""
    return f"{float(value):.1f}%"


def _pnl_class(value: Any) -> str:
    if pd.isna(value):
        return "blank"
    return "pos" if float(value) > 0 else "neg" if float(value) < 0 else "zero"


def _heat_color(value: Any, max_abs: float) -> str:
    if pd.isna(value) or max_abs <= 0:
        return "#f4f6f8"
    number = float(value)
    intensity = min(abs(number) / max_abs, 1.0)
    if number >= 0:
        red = int(238 - 102 * intensity)
        green = int(247 - 88 * intensity)
        blue = int(239 - 101 * intensity)
    else:
        red = int(253 - 67 * intensity)
        green = int(237 - 91 * intensity)
        blue = int(237 - 87 * intensity)
    return f"rgb({red}, {green}, {blue})"


def _dataframe_table(
    df: pd.DataFrame,
    *,
    columns: list[str],
    percent_columns: set[str] | None = None,
    pnl_columns: set[str] | None = None,
    max_rows: int | None = None,
) -> str:
    percent_columns = percent_columns or set()
    pnl_columns = pnl_columns or set()
    view = df.loc[:, [column for column in columns if column in df.columns]].copy()
    if max_rows is not None:
        view = view.head(max_rows)
    header = "".join(f"<th>{escape(column)}</th>" for column in view.columns)
    rows = []
    for _, row in view.iterrows():
        cells = []
        for column in view.columns:
            value = row[column]
            css_class = _pnl_class(value) if column in pnl_columns else ""
            if column in percent_columns:
                text = _format_pct(value)
            elif pd.api.types.is_number(value):
                text = _format_number(value)
            else:
                text = escape(str(value))
            cells.append(f'<td class="{css_class}">{text}</td>')
        rows.append("<tr>" + "".join(cells) + "</tr>")
    return f"<table><thead><tr>{header}</tr></thead><tbody>{''.join(rows)}</tbody></table>"


def _monthly_matrix_html(matrix: pd.DataFrame) -> str:
    if matrix.empty:
        return "<p class=\"muted\">Нет успешных OOS-строк для матрицы.</p>"
    max_abs = float(matrix.abs().max().max())
    header = "<th>ticker / model</th>" + "".join(
        f"<th>{escape(str(column))}</th>" for column in matrix.columns
    )
    body_rows = []
    for row_label, row in matrix.iterrows():
        cells = [f"<th>{escape(str(row_label))}</th>"]
        for value in row:
            cells.append(
                f'<td class="{_pnl_class(value)}" style="background:{_heat_color(value, max_abs)}">'
                f"{_format_number(value)}</td>"
            )
        body_rows.append("<tr>" + "".join(cells) + "</tr>")
    return f"<table class=\"heatmap\"><thead><tr>{header}</tr></thead><tbody>{''.join(body_rows)}</tbody></table>"


def _ticker_equity_svg(summary: pd.DataFrame, ticker: str) -> str:
    ok = ok_rows(summary)
    ok = ok[ok["ticker"] == ticker].copy()
    if ok.empty:
        return ""
    ok = ok.sort_values(["model_dir", "month"])
    months = sorted(ok["month"].unique().tolist())
    if not months:
        return ""

    width = 920
    height = 260
    pad_left = 58
    pad_right = 18
    pad_top = 18
    pad_bottom = 42
    plot_width = width - pad_left - pad_right
    plot_height = height - pad_top - pad_bottom

    series: dict[str, list[float]] = {}
    all_values = [0.0]
    for model_dir, group in ok.groupby("model_dir", sort=True):
        by_month = group.set_index("month")["total_pnl"].to_dict()
        cumulative = []
        running = 0.0
        for month in months:
            running += float(by_month.get(month, 0.0))
            cumulative.append(running)
            all_values.append(running)
        series[model_dir] = cumulative

    min_value = min(all_values)
    max_value = max(all_values)
    if min_value == max_value:
        min_value -= 1
        max_value += 1

    def x_pos(index: int) -> float:
        if len(months) == 1:
            return pad_left + plot_width / 2
        return pad_left + plot_width * index / (len(months) - 1)

    def y_pos(value: float) -> float:
        return pad_top + plot_height * (max_value - value) / (max_value - min_value)

    palette = ["#2454a6", "#15925d", "#b45f06", "#7a3db8", "#c62828", "#00838f", "#455a64", "#ad1457"]
    paths = []
    labels = []
    for i, (model_dir, values) in enumerate(series.items()):
        color = palette[i % len(palette)]
        points = " ".join(f"{x_pos(j):.1f},{y_pos(value):.1f}" for j, value in enumerate(values))
        paths.append(f'<polyline points="{points}" fill="none" stroke="{color}" stroke-width="2.4"/>')
        labels.append(
            f'<span><i style="background:{color}"></i>{escape(model_dir)}</span>'
        )

    zero_y = y_pos(0.0)
    month_labels = "".join(
        f'<text x="{x_pos(i):.1f}" y="{height - 15}" text-anchor="middle">{escape(month)}</text>'
        for i, month in enumerate(months)
    )
    grid = (
        f'<line x1="{pad_left}" x2="{width - pad_right}" y1="{zero_y:.1f}" y2="{zero_y:.1f}" class="zero-line"/>'
        f'<text x="{pad_left - 8}" y="{zero_y + 4:.1f}" text-anchor="end">0</text>'
        f'<text x="{pad_left - 8}" y="{pad_top + 4}" text-anchor="end">{_format_number(max_value)}</text>'
        f'<text x="{pad_left - 8}" y="{height - pad_bottom}" text-anchor="end">{_format_number(min_value)}</text>'
    )
    return (
        f'<div class="svg-wrap"><svg viewBox="0 0 {width} {height}" role="img" '
        f'aria-label="{escape(ticker)} cumulative OOS PnL">'
        f'<rect x="0" y="0" width="{width}" height="{height}" fill="#ffffff"/>'
        f"{grid}{''.join(paths)}{month_labels}</svg>"
        f'<div class="legend">{"".join(labels)}</div></div>'
    )


def build_html(summary: pd.DataFrame, leaderboard: pd.DataFrame, matrix: pd.DataFrame) -> str:
    normalized = normalize_summary(summary)
    ok = ok_rows(normalized)
    errors = error_rows(normalized)
    tickers = sorted(ok["ticker"].unique().tolist())
    months = sorted(ok["month"].unique().tolist())
    best_by_ticker = leaderboard[leaderboard["rank"] == 1].copy()

    cards = [
        ("Тикеров", len(tickers)),
        ("Моделей", ok[["ticker", "model_dir"]].drop_duplicates().shape[0]),
        ("Месяцев", len(months)),
        ("OOS PnL", _format_number(ok["total_pnl"].sum())),
        ("Ошибок", len(errors)),
    ]
    card_html = "".join(
        f"<div class=\"metric\"><span>{escape(label)}</span><strong>{escape(str(value))}</strong></div>"
        for label, value in cards
    )

    top_table = _dataframe_table(
        best_by_ticker,
        columns=["ticker", "model_dir", "months", "total_pnl", "win_months_pct", "trades", "max_drawdown", "score"],
        percent_columns={"win_months_pct"},
        pnl_columns={"total_pnl", "max_drawdown", "score"},
    )
    leaderboard_table = _dataframe_table(
        leaderboard,
        columns=[
            "rank",
            "ticker",
            "model_dir",
            "months",
            "total_pnl",
            "median_pnl_month",
            "win_months_pct",
            "trades",
            "pnl_per_trade",
            "max_drawdown",
            "worst_month",
            "worst_month_pnl",
            "score",
        ],
        percent_columns={"win_months_pct"},
        pnl_columns={"total_pnl", "median_pnl_month", "pnl_per_trade", "max_drawdown", "worst_month_pnl", "score"},
    )

    ticker_sections = []
    for ticker in tickers:
        ticker_leaderboard = leaderboard[leaderboard["ticker"] == ticker]
        ticker_matrix = matrix[matrix.index.to_series().str.startswith(f"{ticker} / ")]
        ticker_sections.append(
            f"<section><h2>{escape(ticker)}</h2>"
            f"{_ticker_equity_svg(normalized, ticker)}"
            f"{_dataframe_table(ticker_leaderboard, columns=['rank', 'model_dir', 'months', 'total_pnl', 'win_months_pct', 'trades', 'max_drawdown', 'score'], percent_columns={'win_months_pct'}, pnl_columns={'total_pnl', 'max_drawdown', 'score'})}"
            f"{_monthly_matrix_html(ticker_matrix)}</section>"
        )

    errors_html = (
        _dataframe_table(
            errors,
            columns=["ticker", "model_dir", "sentiment_model", "month", "error"],
        )
        if not errors.empty
        else "<p class=\"muted\">Ошибок и пропусков нет.</p>"
    )

    return f"""<!doctype html>
<html lang="ru">
<head>
<meta charset="utf-8">
<title>OOS Dashboard</title>
<style>
:root {{
  color-scheme: light;
  --bg: #f4f6f8;
  --panel: #ffffff;
  --text: #17212f;
  --muted: #667085;
  --line: #d9e0e7;
  --pos: #087443;
  --neg: #b42318;
  --zero: #475467;
}}
* {{ box-sizing: border-box; }}
body {{
  margin: 0;
  background: var(--bg);
  color: var(--text);
  font-family: "Segoe UI", Arial, sans-serif;
  font-size: 14px;
}}
main {{ max-width: 1480px; margin: 0 auto; padding: 28px; }}
h1 {{ margin: 0 0 6px; font-size: 30px; }}
h2 {{ margin: 0 0 16px; font-size: 22px; }}
p {{ color: var(--muted); }}
section {{
  background: var(--panel);
  border: 1px solid var(--line);
  border-radius: 8px;
  margin: 18px 0;
  padding: 18px;
  box-shadow: 0 1px 2px rgba(16, 24, 40, 0.04);
}}
.metrics {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 12px; margin: 18px 0; }}
.metric {{ background: var(--panel); border: 1px solid var(--line); border-radius: 8px; padding: 14px 16px; }}
.metric span {{ color: var(--muted); display: block; margin-bottom: 6px; }}
.metric strong {{ font-size: 24px; }}
.table-scroll {{ overflow-x: auto; }}
table {{ border-collapse: collapse; width: 100%; margin: 12px 0; }}
th, td {{ border-bottom: 1px solid var(--line); padding: 8px 10px; text-align: right; white-space: nowrap; }}
th:first-child, td:first-child {{ text-align: left; }}
thead th {{ background: #eef2f6; position: sticky; top: 0; z-index: 1; }}
.heatmap th {{ background: #eef2f6; }}
.pos {{ color: var(--pos); font-weight: 650; }}
.neg {{ color: var(--neg); font-weight: 650; }}
.zero {{ color: var(--zero); }}
.blank {{ color: var(--muted); }}
.muted {{ color: var(--muted); }}
.svg-wrap {{ border: 1px solid var(--line); border-radius: 8px; padding: 10px; margin-bottom: 14px; overflow-x: auto; }}
svg {{ width: 100%; min-width: 760px; height: auto; }}
svg text {{ fill: var(--muted); font-size: 11px; }}
.zero-line {{ stroke: #98a2b3; stroke-width: 1; stroke-dasharray: 4 4; }}
.legend {{ display: flex; flex-wrap: wrap; gap: 10px 18px; color: var(--muted); margin-top: 8px; }}
.legend i {{ display: inline-block; width: 20px; height: 3px; margin-right: 6px; vertical-align: middle; }}
</style>
</head>
<body>
<main>
<h1>OOS Dashboard</h1>
<p>Leave-one-month-out: правила строятся на всех месяцах, кроме тестового; тестовый месяц используется только для симуляции.</p>
<div class="metrics">{card_html}</div>
<section><h2>Лучшие модели по тикерам</h2><div class="table-scroll">{top_table}</div></section>
<section><h2>Leaderboard</h2><div class="table-scroll">{leaderboard_table}</div></section>
<section><h2>Месячная heatmap</h2><div class="table-scroll">{_monthly_matrix_html(matrix)}</div></section>
{''.join(ticker_sections)}
<section><h2>Ошибки и пропуски</h2><div class="table-scroll">{errors_html}</div></section>
</main>
</body>
</html>
"""


def _write_sheet(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame, index: bool = False) -> None:
    df.to_excel(writer, sheet_name=sheet_name, index=index)


def _style_workbook(path: Path) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 10), 32)
        if worksheet.max_row > 1 and worksheet.max_column > 1:
            worksheet.auto_filter.ref = worksheet.dimensions

    for sheet_name in ("Monthly_Matrix",):
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            if worksheet.max_row > 1 and worksheet.max_column > 1:
                cell_range = f"B2:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
                worksheet.conditional_formatting.add(
                    cell_range,
                    ColorScaleRule(
                        start_type="min",
                        start_color="F4B7B2",
                        mid_type="num",
                        mid_value=0,
                        mid_color="FFFFFF",
                        end_type="max",
                        end_color="A9D18E",
                    ),
                )

    workbook.save(path)


def write_excel_report(
    *,
    summary: pd.DataFrame,
    leaderboard: pd.DataFrame,
    matrix: pd.DataFrame,
    output_xlsx: Path,
) -> None:
    normalized = normalize_summary(summary)
    errors = error_rows(normalized)
    dashboard = pd.DataFrame(
        [
            {"metric": "tickers", "value": ok_rows(normalized)["ticker"].nunique()},
            {"metric": "models", "value": ok_rows(normalized)[["ticker", "model_dir"]].drop_duplicates().shape[0]},
            {"metric": "months", "value": ok_rows(normalized)["month"].nunique()},
            {"metric": "total_oos_pnl", "value": ok_rows(normalized)["total_pnl"].sum()},
            {"metric": "errors", "value": len(errors)},
        ]
    )
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        _write_sheet(writer, "Dashboard", dashboard)
        _write_sheet(writer, "Leaderboard", leaderboard)
        _write_sheet(writer, "Monthly_Matrix", matrix.reset_index().rename(columns={"row_label": "ticker_model"}))
        for ticker in sorted(ok_rows(normalized)["ticker"].unique().tolist()):
            ticker_leaderboard = leaderboard[leaderboard["ticker"] == ticker]
            _write_sheet(writer, ticker[:31], ticker_leaderboard)
        _write_sheet(writer, "Raw_Summary", normalized)
        _write_sheet(writer, "Errors", errors)
    _style_workbook(output_xlsx)


def build_report(
    *,
    summary_csv: Path,
    output_html: Path,
    output_xlsx: Path,
) -> None:
    if not summary_csv.exists():
        raise FileNotFoundError(f"summary.csv не найден: {summary_csv}")
    summary = pd.read_csv(summary_csv)
    summary = normalize_summary(summary)
    leaderboard = build_leaderboard(summary)
    matrix = build_monthly_matrix(summary)

    output_html.parent.mkdir(parents=True, exist_ok=True)
    output_html.write_text(build_html(summary, leaderboard, matrix), encoding="utf-8")
    write_excel_report(
        summary=summary,
        leaderboard=leaderboard,
        matrix=matrix,
        output_xlsx=output_xlsx,
    )


@app.command()
def main(
    summary_csv: Path = typer.Option(
        ROOT / "oos" / "results" / "summary.csv",
        "--summary-csv",
        help="Путь к summary.csv, созданному oos.run_oos.",
    ),
    output_html: Path = typer.Option(
        ROOT / "oos" / "results" / "oos_report.html",
        "--output-html",
        help="Путь для HTML dashboard.",
    ),
    output_xlsx: Path = typer.Option(
        ROOT / "oos" / "results" / "oos_report.xlsx",
        "--output-xlsx",
        help="Путь для Excel workbook.",
    ),
) -> None:
    build_report(
        summary_csv=summary_csv,
        output_html=output_html,
        output_xlsx=output_xlsx,
    )
    typer.echo(f"HTML: {output_html}")
    typer.echo(f"Excel: {output_xlsx}")


if __name__ == "__main__":
    app()
